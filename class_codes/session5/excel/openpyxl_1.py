#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook


# In[2]:


book_a = load_workbook("one.xlsx")


# In[3]:


print(book_a.sheetnames)


# In[4]:


sheet_a = book_a['thursday']


# In[5]:


print(sheet_a.cell(row=7, column=1))


# In[6]:


ca = sheet_a.cell(row=7, column=1)


# In[7]:


ca.value


# In[8]:


sheet_a['A7'].value


# In[9]:


b7 = sheet_a['B7']


# In[10]:


print(b7.comment)


# In[11]:


print(b7.comment.author)


# In[12]:


from openpyxl import Workbook


# In[13]:


book_b = Workbook()


# In[14]:


sheet_a = book_b.active    #  sheet1


# In[15]:


sheet_a.title = "batman"


# In[16]:


sheet_a['A3'] = "hello"


# In[17]:


sheet_a.cell(column=2, row=4).value = 55


# In[18]:


book_b.save("two.xlsx")


# In[ ]:




